from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, update_session_auth_hash, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.utils import timezone
import openpyxl
import json
from django.db import IntegrityError
from urllib.parse import quote
from django.db.models import Q
from django.core.exceptions import PermissionDenied
from .models import Notification
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
import requests
from core.fcm import send_fcm_v1  # hoặc từ app_name.fcm nếu bạn để ở nơi khác

from .models import (
    Teacher, Subject, ClassRoom, Semester, 
    TeacherAssignment, Student, GradeType, Grade, GradeHistory
)
from .forms import (
    CustomAuthenticationForm, UserForm, TeacherForm, SubjectForm, 
    ClassroomForm, GradeUploadForm, GradeBulkUpdateForm, SemesterForm, StudentImportForm
)

# Helper functions
def is_admin(user):
    return user.is_superuser

def is_teacher(user):
    return hasattr(user, 'teacher')

# Common views
def home(request):
    return render(request, 'core/home.html')

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if not username or not password:
            messages.error(request, 'Vui lòng nhập đầy đủ tên đăng nhập và mật khẩu')
            return redirect('login')
            
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if user.is_active:
                login(request, user)
                
                # Redirect based on user type
                if user.is_superuser:
                    return redirect('admin_dashboard')
                elif hasattr(user, 'teacher'):
                    return redirect('teacher_dashboard')
                elif hasattr(user, 'student'):
                    return redirect('dashboard_students')
                else:
                    messages.error(request, 'Tài khoản không có quyền truy cập')
                    logout(request)
                    return redirect('login')
            else:
                messages.error(request, 'Tài khoản đã bị khóa')
                return redirect('login')
        else:
            messages.error(request, 'Sai tên đăng nhập hoặc mật khẩu')
            return redirect('login')
            
    return render(request, 'core/login.html')

def logout_view(request):
    logout(request)
    messages.success(request, 'Bạn đã đăng xuất thành công.')
    return redirect('home')

# Admin views
@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    teachers = Teacher.objects.all()
    subjects = Subject.objects.all()
    classrooms = ClassRoom.objects.all()
    semesters = Semester.objects.all()
    assignments = TeacherAssignment.objects.all()
    students = Student.objects.all()
    
    return render(request, 'core/admin/dashboard.html', {
        'teachers': teachers,
        'subjects': subjects,
        'classrooms': classrooms,
        'semesters': semesters,
        'assignments': assignments,
        'students': students
    })

@login_required
@user_passes_test(is_admin)
def teacher_list(request):
    teachers = Teacher.objects.all()
    return render(request, 'core/admin/teacher_list.html', {'teachers': teachers})

@login_required
@user_passes_test(is_admin)
def teacher_create(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        phone = request.POST.get('phone')
        
        # Create user
        user = User.objects.create_user(
            username=username,
            email=email,
            password='password123',  # Default password
            first_name=first_name,
            last_name=last_name
        )
        
        # Create teacher
        teacher = Teacher.objects.create(
            user=user,
            phone=phone,
            is_active=True
        )
        
        messages.success(request, f'Đã tạo tài khoản giáo viên {username} thành công với mật khẩu mặc định "password123"')
        return redirect('teacher_list')
    
    return render(request, 'core/admin/teacher_create.html')

@login_required
@user_passes_test(is_admin)
def teacher_detail(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    teacher_assignments = TeacherAssignment.objects.filter(teacher=teacher)
    
    if request.method == 'POST':
        form = TeacherForm(request.POST, instance=teacher, user=teacher.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Thông tin giáo viên đã được cập nhật.')
            return redirect('teacher_detail', teacher_id=teacher.id)
    else:
        form = TeacherForm(instance=teacher, user=teacher.user)
    
    return render(request, 'core/admin/teacher_detail.html', {
        'teacher': teacher, 
        'form': form,
        'teacher_assignments': teacher_assignments
    })

@login_required
@user_passes_test(is_admin)
def teacher_toggle_status(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    teacher.is_active = not teacher.is_active
    teacher.save()
    
    status = 'kích hoạt' if teacher.is_active else 'khóa'
    messages.success(request, f'Đã {status} tài khoản giáo viên {teacher.user.username}')
    return redirect('teacher_list')

@login_required
@user_passes_test(is_admin)
def teacher_delete(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    username = teacher.user.username
    teacher.user.delete()  # This will also delete the teacher due to the CASCADE
    
    messages.success(request, f'Đã xóa tài khoản giáo viên {username}')
    return redirect('teacher_list')

@login_required
@user_passes_test(is_admin)
def teacher_reset_password(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    if request.method == 'POST':
        new_password = request.POST.get('new_password')
        if new_password:
            teacher.user.set_password(new_password)
            teacher.user.save()
            messages.success(request, f"Password for {teacher.user.get_full_name()} has been reset.")
        return redirect('teacher_list')
    return render(request, 'core/admin/teacher_reset_password.html', {'teacher': teacher})

@login_required
@user_passes_test(is_admin)
def teacher_update(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=teacher.user)
        teacher_form = TeacherForm(request.POST, instance=teacher)
        if user_form.is_valid() and teacher_form.is_valid():
            user_form.save()
            teacher_form.save()
            messages.success(request, f"Teacher {teacher.user.get_full_name()} has been updated successfully.")
            return redirect('teacher_detail', teacher_id=teacher.id)
    else:
        user_form = UserForm(instance=teacher.user)
        teacher_form = TeacherForm(instance=teacher)
    
    return render(request, 'core/admin/teacher_form.html', {
        'user_form': user_form,
        'teacher_form': teacher_form,
        'teacher': teacher,
        'is_update': True
    })

@login_required
@user_passes_test(is_admin)
def subject_list(request):
    subjects = Subject.objects.all()
    return render(request, 'core/admin/subject_list.html', {'subjects': subjects})

@login_required
@user_passes_test(is_admin)
def subject_create(request):
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Môn học mới đã được tạo.')
            return redirect('subject_list')
    else:
        form = SubjectForm()
    
    return render(request, 'core/admin/subject_form.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def subject_edit(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            messages.success(request, 'Môn học đã được cập nhật.')
            return redirect('subject_list')
    else:
        form = SubjectForm(instance=subject)
    return render(request, 'core/admin/subject_form.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def subject_delete(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        subject.delete()
        messages.success(request, 'Môn học đã được xóa.')
        return redirect('subject_list')
    return render(request, 'core/admin/subject_confirm_delete.html', {'subject': subject})

@login_required
@user_passes_test(is_admin)
def classroom_list(request):
    classrooms = ClassRoom.objects.all()
    return render(request, 'core/admin/classroom_list.html', {'classrooms': classrooms})

@login_required
@user_passes_test(is_admin)
def classroom_create(request):
    if request.method == 'POST':
        form = ClassroomForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Thêm lớp thành công")
            return redirect('classroom_list')
    else:
        form = ClassroomForm()
    
    return render(request, 'core/admin/classroom_form.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def classroom_detail(request, classroom_id):
    classroom = get_object_or_404(ClassRoom, id=classroom_id)
    students = classroom.students.all()
    context = {
        'classroom': classroom,
        'students': students,
    }
    return render(request, 'core/admin/classroom_detail.html', context)

@login_required
@user_passes_test(is_admin)
def classroom_import_students(request, classroom_id):
    classroom = get_object_or_404(ClassRoom, id=classroom_id)

    if request.method == 'POST':
        form = StudentImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']

            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active

                headers = [sheet.cell(row=1, column=col).value.lower().strip()
                           for col in range(1, sheet.max_column + 1) if sheet.cell(row=1, column=col).value]

                required_columns = ['student_id', 'name']
                missing_columns = [col for col in required_columns if col not in headers]

                if missing_columns:
                    messages.error(request, f"File Excel thiếu các cột bắt buộc: {', '.join(missing_columns)}")
                    return redirect('classroom_import_students', classroom_id=classroom.id)

                students_created = 0
                duplicate_students = []

                for row in range(2, sheet.max_row + 1):
                    student_id = str(sheet.cell(row=row, column=headers.index('student_id') + 1).value or '').strip()
                    name = str(sheet.cell(row=row, column=headers.index('name') + 1).value or '').strip()

                    email = (str(sheet.cell(row=row, column=headers.index('email') + 1).value) or '').strip() if 'email' in headers else ''
                    phone = (str(sheet.cell(row=row, column=headers.index('phone') + 1).value) or '').strip() if 'phone' in headers else ''

                    if not student_id or not name:
                        continue

                    existing_student = Student.objects.filter(student_id=student_id).first()

                    if existing_student:
                        if classroom.students.filter(id=existing_student.id).exists():
                            # Sinh viên đã thuộc lớp này → không thêm lại
                            duplicate_students.append(f"{existing_student.student_id} - {existing_student.name}")
                            continue
                        else:
                            # Sinh viên đã tồn tại nhưng chưa thuộc lớp → chỉ thêm vào lớp
                            classroom.students.add(existing_student)
                            students_created += 1
                            continue

                    # Nếu chưa tồn tại, thì tạo mới
                    try:
                        student = Student.objects.create(
                            student_id=student_id,
                            name=name,
                            phone=phone
                        )
                        classroom.students.add(student)
                        students_created += 1

                        user = User.objects.create_user(
                            username=student_id,
                            email=email if email else f"{student_id}@example.com",
                            password='password123'
                        )
                        student.user = user
                        student.save()

                    except IntegrityError:
                        duplicate_students.append(f"{student_id} (lỗi tạo trùng)")
                        continue
                    except Exception as e:
                        print(f"Error processing row {row}: {e}")
                        continue

                if students_created > 0:
                    messages.success(request, f"Đã thêm {students_created} sinh viên mới.")
                if duplicate_students:
                    messages.error(request, f"Các sinh viên bị trùng không được thêm: {', '.join(duplicate_students)}.")

            except Exception as e:
                messages.error(request, f"Lỗi khi xử lý file Excel: {str(e)}")

            return redirect('classroom_detail', classroom_id=classroom.id)

    else:
        form = StudentImportForm()

    return render(request, 'core/admin/classroom_import_students.html', {
        'form': form,
        'classroom': classroom
    })



@login_required
@user_passes_test(is_admin)
def teacher_assignment_list(request):
    assignments = TeacherAssignment.objects.all()
    return render(request, 'core/admin/assignment_list.html', {'assignments': assignments})

@login_required
@user_passes_test(is_admin)
def teacher_assignment_create(request):
    teachers = Teacher.objects.filter(is_active=True)
    subjects = Subject.objects.all()
    classrooms = ClassRoom.objects.all()
    semesters = Semester.objects.all()
    
    # Check if there are any semesters before processing the form
    if request.method == 'POST' and semesters.exists():
        teacher_id = request.POST.get('teacher')
        subject_id = request.POST.get('subject')
        classroom_ids = request.POST.getlist('classrooms')
        semester_id = request.POST.get('semester')
        
        teacher = get_object_or_404(Teacher, id=teacher_id)
        subject = get_object_or_404(Subject, id=subject_id)
        semester = get_object_or_404(Semester, id=semester_id)
        
        for classroom_id in classroom_ids:
            classroom = get_object_or_404(ClassRoom, id=classroom_id)
            
            # Check if assignment already exists
            assignment, created = TeacherAssignment.objects.get_or_create(
                teacher=teacher,
                subject=subject,
                classroom=classroom,
                semester=semester
            )
            
            if created:
                # Gửi FCM nếu giáo viên có token
                if teacher.fcm_token:
                    title = "Thông báo phân công lớp"
                    body = f"Lớp {classroom.name} môn {subject.name} kỳ {semester.name}"
                    success, message_id = send_fcm_v1(teacher.fcm_token, title, body)
                    print("✅ Gửi FCM:", success, message_id)
                else:
                    print(f"⚠️ Không tìm thấy fcm_token cho giáo viên {teacher.user.username}")

        

        messages.success(request, 'Phân công giảng dạy đã được tạo.')
        return redirect('teacher_assignment_list')
    
    # If no semesters exist and form is submitted, show an error
    if request.method == 'POST' and not semesters.exists():
        messages.error(request, 'Không thể tạo phân công giảng dạy vì chưa có học kỳ nào. Vui lòng tạo học kỳ trước.')
    
    return render(request, 'core/admin/assignment_create.html', {
        'teachers': teachers,
        'subjects': subjects,
        'classrooms': classrooms,
        'semesters': semesters
    })

@login_required
@user_passes_test(is_admin)
def teacher_assignment_delete(request, assignment_id):
    if request.method == 'POST':
        try:
            assignment = get_object_or_404(TeacherAssignment, id=assignment_id)
            assignment.delete()
            return JsonResponse({'success': True, 'message': 'Phân công giảng dạy đã được xóa thành công.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ.'})

# Teacher views
@login_required
@user_passes_test(is_teacher)
def teacher_dashboard(request):
    teacher = request.user.teacher
    current_date = timezone.now().date()
    
    # Get active semester
    active_semesters = Semester.objects.filter(is_active=True)
    
    # Get all assignments for this teacher
    assignments = TeacherAssignment.objects.filter(teacher=teacher)
    
    if active_semesters.exists():
        # Filter assignments by active semesters if requested
        semester_id = request.GET.get('semester')
        if semester_id:
            try:
                selected_semester = Semester.objects.get(id=semester_id)
                assignments = assignments.filter(semester=selected_semester)
            except Semester.DoesNotExist:
                pass
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        assignments = assignments.filter(subject__name__icontains=search_query)
    
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')[:5]
    unread_count = Notification.objects.filter(user=request.user, is_read=False).count()

    return render(request, 'core/teacher/dashboard.html', {
        'teacher': teacher,
        'assignments': assignments,
        'active_semesters': active_semesters,
        'current_date': current_date,
        'search_query': search_query,
        'notifications': notifications,
        'unread_count': unread_count
    })
    

@login_required
@user_passes_test(is_teacher)
def teacher_profile(request):
    teacher = request.user.teacher
    
    if request.method == 'POST':
        form = TeacherForm(request.POST, instance=teacher, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Thông tin cá nhân đã được cập nhật.')
            return redirect('teacher_profile')
    else:
        form = TeacherForm(instance=teacher, user=request.user)
    
    return render(request, 'core/teacher/profile.html', {'form': form, 'teacher': teacher})

@login_required
@user_passes_test(is_teacher)
def teacher_change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Mật khẩu của bạn đã được cập nhật!')
            return redirect('teacher_profile')
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'core/teacher/change_password.html', {'form': form})

@login_required
@user_passes_test(is_teacher)
def class_grades(request, assignment_id):
    assignment = get_object_or_404(TeacherAssignment, id=assignment_id)
    
    # Check if the teacher is assigned to this classroom and subject
    if request.user.teacher != assignment.teacher:
        messages.error(request, 'Bạn không có quyền truy cập vào trang này.')
        return redirect('teacher_dashboard')
    
    # Get all students in the classroom
    students = assignment.classroom.students.all().order_by('student_id')
    
    # Get all grade types
    grade_types = GradeType.objects.all().order_by('id')
    
    # Create structured dictionary for all students and grade types
    student_grades = {}
    for student in students:
        student_grades[student.id] = {
            'student': student,
            'grades': {}
        }
        # Initialize with empty structures for all grade types
        for grade_type in grade_types:
            student_grades[student.id]['grades'][grade_type.id] = {
                'id': None,
                'value': None,
                'created_at': None,
                'updated_at': None
            }
    
    # Get all grades for this assignment with optimized query
    grades = Grade.objects.filter(
        teacher_assignment=assignment
    ).select_related('student', 'grade_type')
    
    # Populate the student_grades dictionary with actual grades
    for grade in grades:
        if grade.student_id in student_grades and grade.grade_type_id in student_grades[grade.student_id]['grades']:
            # Ensure value is always a string to prevent JavaScript conversion issues
            grade_value = str(grade.value) if grade.value is not None else ''
            
            student_grades[grade.student_id]['grades'][grade.grade_type_id] = {
                'id': grade.id,
                'value': grade_value,
                'created_at': grade.created_at,
                'updated_at': grade.updated_at
            }
    
    upload_form = GradeUploadForm()
    
    context = {
        'assignment': assignment,
        'students': students,
        'grade_types': grade_types,
        'student_grades': student_grades,
        'upload_form': upload_form,
    }
    return render(request, 'core/teacher/class_grades.html', context)

@login_required
@user_passes_test(is_teacher)
def teacher_class_students(request, assignment_id):
    assignment = get_object_or_404(TeacherAssignment, id=assignment_id)
    
    # Check if the teacher is authorized to view this assignment
    if assignment.teacher.user != request.user:
        return HttpResponseForbidden("Bạn không có quyền truy cập trang này.")
    
    students = assignment.classroom.students.all().order_by('name')
    
    return render(request, 'core/teacher/class_students.html', {
        'assignment': assignment,
        'students': students
    })

@login_required
@user_passes_test(is_teacher)
def update_grade(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            student_id = data.get('student_id')
            grade_type_id = data.get('grade_type_id')
            grade_id = data.get('grade_id')
            new_value = data.get('value')
            version = data.get('version')

            student = get_object_or_404(Student, id=student_id)
            grade_type = get_object_or_404(GradeType, id=grade_type_id)

            if grade_id:
                grade = get_object_or_404(Grade, id=grade_id)
                if grade.value != new_value:
                    # Record grade history
                    GradeHistory.objects.create(
                        grade=grade,
                        old_value=grade.value,
                        new_value=new_value,
                        modified_by=request.user,
                        action='update'
                    )
                    grade.value = new_value
                    grade.last_modified_by = request.user
                    grade.save()
            else:
                grade = Grade.objects.create(
                    student=student,
                    teacher_assignment=student.classroom.teacherassignment_set.get(teacher=request.user.teacher),
                    grade_type=grade_type,
                    value=new_value,
                    last_modified_by=request.user
                )
                # Record grade history for new grade
                GradeHistory.objects.create(
                    grade=grade,
                    new_value=new_value,
                    modified_by=request.user,
                    action='create'
                )

            # Calculate average grade
            all_grades = Grade.objects.filter(
                student=student,
                teacher_assignment=grade.teacher_assignment
            ).values_list('value', flat=True)
            
            average = sum(all_grades) / len(all_grades) if all_grades else None

            return JsonResponse({
                'success': True,
                'grade_id': grade.id,
                'version': version,
                'average': round(average, 2) if average is not None else None
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
@user_passes_test(is_teacher)
def bulk_update_grades(request, assignment_id):
    assignment = get_object_or_404(TeacherAssignment, id=assignment_id)
    
    # Check if the teacher is authorized to update grades for this assignment
    if assignment.teacher.user != request.user:
        return HttpResponseForbidden("Bạn không có quyền cập nhật điểm cho lớp này.")
    
    if request.method == 'POST':
        form = GradeBulkUpdateForm(request.POST, classroom=assignment.classroom)
        if form.is_valid():
            grade_value = form.cleaned_data['grade_value']
            students = form.cleaned_data['students']
            grade_type_id = request.POST.get('grade_type')
            
            grade_type = get_object_or_404(GradeType, id=grade_type_id)
            
            for student in students:
                # Get or create grade
                grade, created = Grade.objects.get_or_create(
                    student=student,
                    teacher_assignment=assignment,
                    grade_type=grade_type,
                    defaults={'value': grade_value, 'last_modified_by': request.user}
                )
                
                if not created:
                    # Record grade history
                    GradeHistory.objects.create(
                        grade=grade,
                        old_value=grade.value,
                        new_value=grade_value,
                        modified_by=request.user
                    )
                    
                    # Update grade
                    grade.value = grade_value
                    grade.last_modified_by = request.user
                    grade.save()
            
            messages.success(request, f'Đã cập nhật điểm {grade_type.name} cho {len(students)} học sinh')
            return redirect('class_grades', assignment_id=assignment.id)
    
    return redirect('class_grades', assignment_id=assignment.id)

@login_required
@user_passes_test(is_teacher)
def upload_grades(request, assignment_id):
    assignment = get_object_or_404(TeacherAssignment, id=assignment_id)
    
    # Check if the teacher is authorized to update grades for this assignment
    if assignment.teacher.user != request.user:
        return HttpResponseForbidden("Bạn không có quyền cập nhật điểm cho lớp này.")
    
    if request.method == 'POST':
        form = GradeUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                
                # Get headers (first row)
                headers = []
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=1, column=col).value
                    if cell_value:
                        # Clean up header name by stripping whitespace and converting to lowercase for comparison
                        headers.append({
                            'original': str(cell_value).strip(), 
                            'normalized': str(cell_value).strip().lower()
                        })
                
                # Extract normalized header names for validation
                normalized_headers = [h['normalized'] for h in headers]
                
                # Validate headers (must include student_id and at least one grade type)
                if 'student_id' not in normalized_headers:
                    messages.error(request, "File Excel phải có cột 'student_id'")
                    return redirect('class_grades', assignment_id=assignment.id)
                
                if len(headers) < 2:
                    messages.error(request, "File Excel phải có ít nhất một cột điểm ngoài cột student_id")
                    return redirect('class_grades', assignment_id=assignment.id)
                
                # Find student_id column index
                student_id_col_idx = None
                for idx, header in enumerate(headers):
                    if header['normalized'] == 'student_id':
                        student_id_col_idx = idx + 1
                        break
                
                # Get all existing grade types for better matching
                existing_grade_types = {gt.name.lower(): gt for gt in GradeType.objects.all()}
                
                # Map normalized header names to actual GradeType objects
                grade_types = {}
                new_grade_types = []
                
                for header in headers:
                    if header['normalized'] != 'student_id':
                        # Check for existing grade type with same name (case insensitive)
                        if header['normalized'] in existing_grade_types:
                            # Use existing grade type
                            grade_types[header['original']] = existing_grade_types[header['normalized']]
                        else:
                            # Create new grade type with original case
                            grade_type = GradeType.objects.create(name=header['original'])
                            grade_types[header['original']] = grade_type
                            new_grade_types.append(header['original'])
                
                # Notify about new grade types created
                if new_grade_types:
                    messages.info(request, f"Đã tạo mới các loại điểm: {', '.join(new_grade_types)}")
                
                # Process data rows
                updated_count = 0
                error_count = 0
                skipped_count = 0
                student_count = 0
                
                # Map students to IDs for faster lookups - make IDs lowercase for case-insensitive matching
                student_map = {student.student_id.lower(): student for student in assignment.classroom.students.all()}
                
                # Track processed student IDs to report statistics
                processed_students = set()
                
                # Process each row in the Excel sheet
                for row in range(2, sheet.max_row + 1):
                    # Get student_id value from the correct column
                    student_id_cell = sheet.cell(row=row, column=student_id_col_idx)
                    if not student_id_cell.value:
                        skipped_count += 1
                        continue
                    
                    # Make sure student_id is a string and normalized for comparison
                    student_id = str(student_id_cell.value).strip()
                    student_id_lower = student_id.lower()
                    
                    # Check if this student exists in the map
                    if student_id_lower in student_map:
                        student = student_map[student_id_lower]
                        processed_students.add(student_id_lower)
                        
                        # Process each grade column
                        for idx, header in enumerate(headers):
                            if header['normalized'] != 'student_id':
                                col_idx = idx + 1
                                cell_value = sheet.cell(row=row, column=col_idx).value
                                
                                if cell_value is not None:
                                    try:
                                        grade_value = float(cell_value)
                                        if 0 <= grade_value <= 10:
                                            grade_type = grade_types[header['original']]
                                            
                                            # Get or create grade
                                            grade, created = Grade.objects.get_or_create(
                                                student=student,
                                                teacher_assignment=assignment,
                                                grade_type=grade_type,
                                                defaults={'value': grade_value, 'last_modified_by': request.user}
                                            )
                                            
                                            if not created:
                                                # Record grade history
                                                GradeHistory.objects.create(
                                                    grade=grade,
                                                    old_value=grade.value,
                                                    new_value=grade_value,
                                                    modified_by=request.user
                                                )
                                                
                                                # Update grade
                                                grade.value = grade_value
                                                grade.last_modified_by = request.user
                                                grade.save()
                                            
                                            updated_count += 1
                                        else:
                                            error_count += 1  # Value out of range
                                    except (ValueError, TypeError):
                                        error_count += 1  # Not a number
                    else:
                        # Student ID not found in this class
                        error_count += 1
                
                student_count = len(processed_students)
                
                if updated_count > 0:
                    messages.success(request, f'Đã cập nhật {updated_count} điểm cho {student_count} sinh viên thành công.')
                
                if error_count > 0:
                    messages.warning(request, f'Có {error_count} giá trị điểm không hợp lệ hoặc mã sinh viên không tồn tại.')
                
                if skipped_count > 0:
                    messages.info(request, f'Đã bỏ qua {skipped_count} dòng không có mã sinh viên.')
                
                return redirect('class_grades', assignment_id=assignment.id)
            
            except Exception as e:
                messages.error(request, f"Lỗi khi xử lý file Excel: {str(e)}")
                return redirect('class_grades', assignment_id=assignment.id)
    else:
        form = GradeUploadForm()
    
    return render(request, 'core/teacher/upload_grades.html', {
        'form': form,
        'assignment': assignment
    })

@login_required
@user_passes_test(is_teacher)
def manage_grade_types(request, assignment_id):
    assignment = get_object_or_404(TeacherAssignment, id=assignment_id)
    
    # Check if the teacher is authorized to manage grade types for this assignment
    if assignment.teacher.user != request.user:
        return HttpResponseForbidden("Bạn không có quyền quản lý loại điểm cho lớp này.")
    
    # Get all grade types
    grade_types = GradeType.objects.all()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create':
            name = request.POST.get('name', '').strip()
            if name:
                grade_type, created = GradeType.objects.get_or_create(name=name)
                if created:
                    messages.success(request, f'Đã tạo loại điểm "{name}" thành công.')
                else:
                    messages.info(request, f'Loại điểm "{name}" đã tồn tại.')
            else:
                messages.error(request, 'Tên loại điểm không được để trống.')
        
        elif action == 'edit':
            grade_type_id = request.POST.get('grade_type_id')
            new_name = request.POST.get('new_name', '').strip()
            
            if grade_type_id and new_name:
                try:
                    grade_type = GradeType.objects.get(id=grade_type_id)
                    old_name = grade_type.name
                    
                    # Check if the new name already exists
                    if GradeType.objects.filter(name=new_name).exclude(id=grade_type_id).exists():
                        messages.error(request, f'Loại điểm "{new_name}" đã tồn tại.')
                    else:
                        grade_type.name = new_name
                        grade_type.save()
                        messages.success(request, f'Đã đổi tên loại điểm "{old_name}" thành "{new_name}".')
                except GradeType.DoesNotExist:
                    messages.error(request, 'Loại điểm không tồn tại.')
            else:
                messages.error(request, 'Thông tin không hợp lệ.')
        
        elif action == 'delete':
            grade_type_id = request.POST.get('grade_type_id')
            
            if grade_type_id:
                try:
                    grade_type = GradeType.objects.get(id=grade_type_id)
                    grade_type_name = grade_type.name
                    
                    # Delete all associated grades first
                    Grade.objects.filter(grade_type=grade_type, teacher_assignment=assignment).delete()
                    
                    # Then delete the grade type
                    grade_type.delete()
                    messages.success(request, f'Đã xóa loại điểm "{grade_type_name}" và tất cả điểm số liên quan thành công.')
                except GradeType.DoesNotExist:
                    messages.error(request, 'Loại điểm không tồn tại.')
            else:
                messages.error(request, 'Thông tin không hợp lệ.')
    
    return render(request, 'core/teacher/manage_grade_types.html', {
        'assignment': assignment,
        'grade_types': grade_types
    })

@login_required
@user_passes_test(is_teacher)
def delete_grade(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            student_id = data.get('student_id')
            grade_type_id = data.get('grade_type_id')

            student = get_object_or_404(Student, id=student_id)
            grade = get_object_or_404(
                Grade,
                student=student,
                grade_type_id=grade_type_id,
                teacher_assignment__teacher=request.user.teacher
            )

            # Record grade history before deletion
            GradeHistory.objects.create(
                grade=grade,
                old_value=grade.value,
                modified_by=request.user,
                action='delete'
            )

            grade.delete()

            return JsonResponse({
                'success': True
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
@user_passes_test(is_teacher)
def grade_history(request):
    if request.method == 'GET':
        student_id = request.GET.get('student_id')
        assignment_id = request.GET.get('assignment_id')
        
        if not student_id or not assignment_id:
            return JsonResponse({'success': False, 'error': 'Thiếu thông tin bắt buộc'})
        
        try:
            student = get_object_or_404(Student, id=student_id)
            assignment = get_object_or_404(TeacherAssignment, id=assignment_id)
            
            # Check if the teacher is authorized
            if assignment.teacher.user != request.user:
                return JsonResponse({'success': False, 'error': 'Bạn không có quyền xem lịch sử điểm cho lớp này'})
            
            # Get all grades for this student in this assignment
            grades = Grade.objects.filter(student=student, teacher_assignment=assignment)
            
            # Get history for all these grades
            history_entries = GradeHistory.objects.filter(
                grade__in=grades
            ).order_by('-modified_at')
            
            if not history_entries:
                return JsonResponse({'success': True, 'html': '<div class="alert alert-info">Không có lịch sử điểm.</div>'})
            
            # Prepare HTML table for history
            html = '<div class="table-responsive">'
            html += '<table class="table table-striped">'
            html += '<thead class="table-light">'
            html += '<tr><th>Thời gian</th><th>Loại điểm</th><th>Điểm cũ</th><th>Điểm mới</th><th>Hành động</th><th>Người thực hiện</th></tr>'
            html += '</thead><tbody>'
            
            for entry in history_entries:
                action_text = ''
                if entry.action == 'update':
                    action_text = 'Cập nhật'
                elif entry.action == 'create':
                    action_text = 'Tạo mới'
                elif entry.action == 'delete':
                    action_text = 'Xóa'
                    
                # Format old and new values
                old_value = entry.old_value if entry.old_value is not None else '-'
                new_value = entry.new_value if entry.new_value is not None else '-'
                
                html += f'<tr>'
                html += f'<td>{entry.modified_at.strftime("%d/%m/%Y %H:%M:%S")}</td>'
                html += f'<td>{entry.grade.grade_type.name}</td>'
                html += f'<td>{old_value}</td>'
                html += f'<td>{new_value}</td>'
                html += f'<td>{action_text}</td>'
                html += f'<td>{entry.modified_by.get_full_name() or entry.modified_by.username}</td>'
                html += f'</tr>'
            
            html += '</tbody></table></div>'
            
            return JsonResponse({'success': True, 'html': html})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Phương thức không hợp lệ'})

@login_required
@user_passes_test(is_admin)
def semester_list(request):
    semesters = Semester.objects.all().order_by('-is_active', 'name')
    return render(request, 'core/admin/semester_list.html', {'semesters': semesters})

@login_required
@user_passes_test(is_admin)
def semester_create(request):
    if request.method == 'POST':
        form = SemesterForm(request.POST)
        if form.is_valid():
            semester = form.save()
            messages.success(request, f'Học kỳ "{semester.name}" đã được tạo thành công.')
            return redirect('semester_list')
    else:
        form = SemesterForm()
    
    return render(request, 'core/admin/semester_form.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def semester_update(request, semester_id):
    semester = get_object_or_404(Semester, id=semester_id)
    
    if request.method == 'POST':
        form = SemesterForm(request.POST, instance=semester)
        if form.is_valid():
            form.save()
            messages.success(request, f'Học kỳ "{semester.name}" đã được cập nhật thành công.')
            return redirect('semester_list')
    else:
        form = SemesterForm(instance=semester)
    
    return render(request, 'core/admin/semester_form.html', {'form': form, 'semester': semester})

@login_required
@user_passes_test(is_admin)
def semester_toggle_status(request, semester_id):
    semester = get_object_or_404(Semester, id=semester_id)
    semester.is_active = not semester.is_active
    semester.save()
    
    status = "kích hoạt" if semester.is_active else "vô hiệu hóa"
    messages.success(request, f'Học kỳ "{semester.name}" đã được {status}.')
    return redirect('semester_list')

@login_required
@user_passes_test(is_admin)
def semester_delete(request, semester_id):
    semester = get_object_or_404(Semester, id=semester_id)
    
    # Check if there are any related assignments
    if TeacherAssignment.objects.filter(semester=semester).exists():
        messages.error(request, f'Không thể xóa học kỳ "{semester.name}" vì đã có phân công giảng dạy liên quan.')
        return redirect('semester_list')
    
    semester_name = semester.name
    semester.delete()
    messages.success(request, f'Học kỳ "{semester_name}" đã được xóa.')
    return redirect('semester_list')

@login_required
@user_passes_test(is_teacher)
def update_grades(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            grades_data = data.get('grades', [])
            
            if not grades_data:
                return JsonResponse({'success': False, 'error': 'Không có dữ liệu điểm để cập nhật'})
            
            # Get the first grade to verify teacher assignment
            first_grade = grades_data[0]
            student = get_object_or_404(Student, id=first_grade['student_id'])
            assignment = get_object_or_404(
                TeacherAssignment,
                classroom=student.classroom,
                teacher=request.user.teacher
            )
            
            updated_grades = []
            for grade_data in grades_data:
                student_id = grade_data['student_id']
                grade_type_id = grade_data['grade_type_id']
                grade_id = grade_data['grade_id']
                value = grade_data['value']
                version = grade_data['version']
                
                # Skip empty values
                if value == '':
                    continue
                
                # Validate value
                try:
                    float_value = float(value)
                    if float_value < 0 or float_value > 10:
                        return JsonResponse({
                            'success': False,
                            'error': f'Điểm phải là số từ 0 đến 10'
                        })
                except ValueError:
                    return JsonResponse({
                        'success': False,
                        'error': f'Điểm không hợp lệ: {value}'
                    })
                
                student = get_object_or_404(Student, id=student_id)
                grade_type = get_object_or_404(GradeType, id=grade_type_id)
                
                if grade_id:
                    grade = get_object_or_404(Grade, id=grade_id)
                    if str(grade.value) != str(value):
                        # Record grade history
                        GradeHistory.objects.create(
                            grade=grade,
                            old_value=grade.value,
                            new_value=value,
                            modified_by=request.user,
                            action='update'
                        )
                        grade.value = float_value
                        grade.last_modified_by = request.user
                        grade.save()
                else:
                    grade = Grade.objects.create(
                        student=student,
                        teacher_assignment=assignment,
                        grade_type=grade_type,
                        value=float_value,
                        last_modified_by=request.user
                    )
                    # Record grade history for new grade
                    GradeHistory.objects.create(
                        grade=grade,
                        new_value=value,
                        modified_by=request.user,
                        action='create'
                    )
                
                updated_grades.append({
                    'grade_type_id': grade_type_id,
                    'grade_id': grade.id,
                    'version': version
                })

            return JsonResponse({
                'success': True,
                'grades': updated_grades
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
@user_passes_test(is_teacher)
def download_grades(request, assignment_id):
    assignment = get_object_or_404(TeacherAssignment, id=assignment_id)
    
    # Check if the teacher is authorized to download grades for this assignment
    if assignment.teacher.user != request.user:
        return HttpResponseForbidden("Bạn không có quyền tải xuống điểm của lớp này.")
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    # Set headers
    headers = ['Mã SV', 'Họ tên']
    grade_types = GradeType.objects.all().order_by('name')
    for grade_type in grade_types:
        headers.append(grade_type.name)
    
    # Write headers
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)
    
    # Get all students in the class
    students = assignment.classroom.students.all().order_by('student_id')
    
    # Write student data
    for row, student in enumerate(students, 2):
        sheet.cell(row=row, column=1, value=student.student_id)
        sheet.cell(row=row, column=2, value=student.name)
        
        # Write grades for each grade type
        for col, grade_type in enumerate(grade_types, 3):
            grade = Grade.objects.filter(
                student=student,
                grade_type=grade_type,
                teacher_assignment=assignment
            ).first()
            sheet.cell(row=row, column=col, value=grade.value if grade else '')
    
    # Generate filename with Vietnamese characters
    filename = f"{assignment.classroom.name}_{assignment.subject.name}.xlsx"
    encoded_filename = quote(filename)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response

@login_required
@user_passes_test(is_admin)
def student_list(request):
    # Get all classrooms for the filter dropdown
    classrooms = ClassRoom.objects.all().order_by('name')
    
    # Get the selected classroom filter
    classroom_id = request.GET.get('classroom')
    search_query = request.GET.get('search', '').strip()
    
    # Get sort parameters
    sort_by = request.GET.get('sort', 'name')  # Default sort by name
    sort_direction = request.GET.get('direction', 'asc')  # Default ascending
    
    # Start with all students
    students = Student.objects.prefetch_related('classrooms').all()
    
    # Apply classroom filter if selected
    if classroom_id:
        students = students.filter(classroom_id=classroom_id)
    
    # Apply search filter if provided
    if search_query:
        students = students.filter(
            Q(student_id__icontains=search_query) |
            Q(name__icontains=search_query)
        )
    
    # Apply sorting
    if sort_by == 'student_id':
        order_by = 'student_id'
    elif sort_by == 'name':
        order_by = 'name'
    elif sort_by == 'classroom':
        order_by = 'classroom__name'
    else:
        order_by = 'name'
    
    # Apply sort direction
    if sort_direction == 'desc':
        order_by = f'-{order_by}'
    
    students = students.order_by(order_by)
    
    context = {
        'students': students,
        'classrooms': classrooms,
        'search_query': search_query,
        'current_sort': sort_by,
        'current_direction': sort_direction,
        'classroom_id': classroom_id
    }
    
    return render(request, 'core/admin/student_list.html', context)

def student_detail(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    grades = Grade.objects.filter(student=student).order_by('-created_at')
    
    context = {
        'student': student,
        'grades': grades,
    }
    return render(request, 'core/admin/student_detail.html', context)

@login_required
@user_passes_test(is_admin)
def student_edit(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == 'POST':
        # Get form data
        student_id_new = request.POST.get('student_id')
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        
        # Validate form data
        form_errors = {}
        
        # Check required fields
        if not student_id_new:
            form_errors['student_id'] = ['Mã sinh viên là bắt buộc']
        if not name:
            form_errors['name'] = ['Họ và tên là bắt buộc']
            
        # Check if student_id is unique (excluding current student)
        if student_id_new != student.student_id and Student.objects.filter(student_id=student_id_new).exists():
            form_errors['student_id'] = ['Mã sinh viên đã tồn tại']
            
        # Validate phone number if provided
        if phone and not phone.isdigit():
            form_errors['phone'] = ['Số điện thoại chỉ được chứa các chữ số']
        elif phone and len(phone) != 10:
            form_errors['phone'] = ['Số điện thoại phải có 10 chữ số']
            
        # Validate email if provided
        if email and '@' not in email:
            form_errors['email'] = ['Email không hợp lệ']
            
        # If there are no errors, save the changes
        if not form_errors:
            # Update student information
            student.student_id = student_id_new
            student.name = name
            student.phone = phone
            
            # Handle user account
            if email:
                # Try to get existing user
                user = None
                if student.user:
                    user = student.user
                    user.email = email
                    user.username = student_id_new
                    user.save()
                else:
                    # Create new user with default password
                    user = User.objects.create_user(
                        username=student_id_new,
                        email=email,
                        password='password123'  # Changed to password123
                    )
                    student.user = user
            elif student.user:
                # If email is removed, delete the user account
                student.user.delete()
                student.user = None
            
            student.save()
            messages.success(request, 'Thông tin sinh viên đã được cập nhật thành công')
            return redirect('student_detail', student_id=student.id)
            
        # If there are errors, re-render form with errors
        context = {
            'student': student,
            'form_errors': form_errors
        }
        return render(request, 'core/admin/student_edit.html', context)
        
    # For GET request, display form with current values
    context = {
        'student': {
            'id': student.id,
            'student_id': student.student_id,
            'name': student.name,
            'email': student.user.email if student.user else '',
            'phone': student.phone or ''
        }
    }
    return render(request, 'core/admin/student_edit.html', context)

@login_required
@user_passes_test(is_admin)
def student_create(request):
    if request.method == 'POST':
        try:
            student_id = request.POST.get('student_id')
            name = request.POST.get('name')
            email = request.POST.get('email')
            phone_number = request.POST.get('phone_number')
            
            # Validate required fields
            if not student_id:
                messages.error(request, 'Vui lòng nhập mã sinh viên')
                return render(request, 'core/admin/student_create.html')
                
            if not name:
                messages.error(request, 'Vui lòng nhập họ tên sinh viên')
                return render(request, 'core/admin/student_create.html')
            
            # Check if student_id already exists in Student model
            if Student.objects.filter(student_id=student_id).exists():
                messages.error(request, f'Mã sinh viên {student_id} đã tồn tại trong hệ thống')
                return render(request, 'core/admin/student_create.html')
                
            # Check if student_id already exists as a username
            if User.objects.filter(username=student_id).exists():
                messages.error(request, f'Mã sinh viên {student_id} đã được sử dụng làm tên đăng nhập')
                return render(request, 'core/admin/student_create.html')
            
            # Create user account first
            user = User.objects.create_user(
                username=student_id,
                password='password123',
                email=email if email else f"{student_id}@example.com"
            )
            user.is_active = True
            user.save()
            
            # Then create student record
            student = Student.objects.create(
                user=user,
                student_id=student_id,
                name=name,
                phone=phone_number if phone_number else None
            )
            
            messages.success(request, f'Đã tạo tài khoản sinh viên {name} (Mã SV: {student_id}) thành công')
            return redirect('student_detail', student_id=student.id)
            
        except Exception as e:
            # If something goes wrong, delete the user if it was created
            if 'user' in locals():
                user.delete()
            messages.error(request, f'Lỗi khi tạo tài khoản sinh viên: {str(e)}')
            return render(request, 'core/admin/student_create.html')
    
    return render(request, 'core/admin/student_create.html')

@login_required
@user_passes_test(is_admin)
def student_delete(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == 'POST':
        # Store student info for message
        student_name = student.name
        student_id_str = student.student_id
        
        # Delete associated user account if exists
        if student.user:
            student.user.delete()
        
        # Delete student
        student.delete()
        
        messages.success(request, f'Đã xóa sinh viên {student_name} (Mã SV: {student_id_str})')
        return redirect('student_list')
    
    return render(request, 'core/admin/student_confirm_delete.html', {'student': student})

@login_required
def student_profile(request):
    if not hasattr(request.user, 'student'):
        raise PermissionDenied
        
    student = request.user.student
    form_errors = {}
    
    if request.method == 'POST':
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        
        # Validate email
        if not email:
            form_errors['email'] = ['Email là bắt buộc']
        
        # Validate phone (optional)
        if phone and not phone.isdigit():
            form_errors['phone'] = ['Số điện thoại chỉ được chứa chữ số']
        elif phone and len(phone) != 10:
            form_errors['phone'] = ['Số điện thoại phải có 10 chữ số']
            
        if not form_errors:
            # Đồng bộ email với user.email
            student.user.email = email
            student.user.save()
            student.phone = phone
            student.save()
            messages.success(request, 'Cập nhật thông tin thành công')
            return redirect('student_profile')
            
    return render(request, 'core/student/profile.html', {
        'student': student,
        'form_errors': form_errors
    })

@login_required
def student_grades(request):
    if not hasattr(request.user, 'student'):
        raise PermissionDenied
        
    student = request.user.student
    
    # Get all teacher assignments for classes the student is enrolled in
    teacher_assignments = TeacherAssignment.objects.filter(
        classroom__in=student.classrooms.all()
    ).select_related(
        'subject',
        'semester',
        'teacher',
        'classroom'
    ).order_by(
        'semester__name',
        'subject__name'
    )
    
    # Get all grade types
    grade_types = GradeType.objects.all().order_by('name')
    
    # Get all grades for this student
    grades = Grade.objects.filter(
        student=student
    ).select_related(
        'teacher_assignment',
        'grade_type'
    )
    
    # Create a dictionary to store grades by assignment and type
    grade_dict = {}
    for grade in grades:
        assignment_id = grade.teacher_assignment_id
        if assignment_id not in grade_dict:
            grade_dict[assignment_id] = {}
        grade_dict[assignment_id][grade.grade_type_id] = grade
    
    # Group assignments by semester
    grouped_subjects = {}
    for assignment in teacher_assignments:
        semester = assignment.semester
        if semester not in grouped_subjects:
            grouped_subjects[semester] = []
            
        # Get grades for each grade type
        subject_grades = []
        total_grade = 0
        grade_count = 0
        
        for grade_type in grade_types:
            grade = grade_dict.get(assignment.id, {}).get(grade_type.id)
            if grade and grade.value is not None:
                total_grade += grade.value
                grade_count += 1
                
            subject_grades.append({
                'type': grade_type.name,
                'value': grade.value if grade else None,
                'updated_at': grade.updated_at if grade else None
            })
        
        # Calculate final grade
        final_grade = round(total_grade / grade_count, 1) if grade_count > 0 else None
        
        # Convert to letter grade
        letter_grade = None
        if final_grade is not None:
            if final_grade >= 8.5:
                letter_grade = 'A'
            elif final_grade >= 7.0:
                letter_grade = 'B'
            elif final_grade >= 5.5:
                letter_grade = 'C'
            elif final_grade >= 4.0:
                letter_grade = 'D'
            else:
                letter_grade = 'F'
            
        grouped_subjects[semester].append({
            'subject': assignment.subject,
            'teacher': assignment.teacher,
            'classroom': assignment.classroom,
            'grades': subject_grades,
            'final_grade': final_grade,
            'letter_grade': letter_grade
        })
    
    context = {
        'student': student,
        'grouped_subjects': grouped_subjects,
        'grade_types': grade_types
    }
    
    return render(request, 'core/student/grades.html', context)

@login_required
def student_change_password(request):
    if not hasattr(request.user, 'student'):
        raise PermissionDenied
        
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        if not request.user.check_password(old_password):
            messages.error(request, 'Mật khẩu cũ không chính xác')
        elif new_password != confirm_password:
            messages.error(request, 'Mật khẩu mới không khớp')
        elif len(new_password) < 8:
            messages.error(request, 'Mật khẩu mới phải có ít nhất 8 ký tự')
        else:
            request.user.set_password(new_password)
            request.user.save()
            messages.success(request, 'Đổi mật khẩu thành công')
            return redirect('login')
            
    return render(request, 'core/student/change_password_students.html', {
        'student': request.user.student
    })

@login_required
def dashboard_students(request):
    if not hasattr(request.user, 'student'):
        raise PermissionDenied
        
    student = request.user.student
    
    # Get all grades for this student
    grades = Grade.objects.filter(student=student)
    
    # Calculate statistics
    total_subjects = TeacherAssignment.objects.filter(classroom__in=student.classrooms.all()).count()
    completed_subjects = grades.values('teacher_assignment').distinct().count()
    ongoing_subjects = total_subjects - completed_subjects
    
    # Calculate GPA
    if grades.exists():
        total_grade = sum(grade.value for grade in grades if grade.value is not None)
        total_count = grades.filter(value__isnull=False).count()
        gpa = round(total_grade / total_count, 2) if total_count > 0 else None
    else:
        gpa = None
    
    context = {
        'student': student,
        'total_subjects': total_subjects,
        'completed_subjects': completed_subjects,
        'ongoing_subjects': ongoing_subjects,
        'gpa': gpa
    }
    
    return render(request, 'core/student/dashboard_students.html', context)

@login_required
@user_passes_test(is_admin)
def classroom_delete(request, classroom_id):
    classroom = get_object_or_404(ClassRoom, id=classroom_id)
    
    if request.method == 'POST':
        # Check if there are any teacher assignments
        if TeacherAssignment.objects.filter(classroom=classroom).exists():
            messages.error(request, f'Không thể xóa lớp "{classroom.name}" vì đã có phân công giảng dạy.')
            return redirect('classroom_list')
        
        # Delete all students in the class
        students = classroom.students.all()
        for student in students:
            if student.user:
                student.user.delete()  # Delete associated user account
            student.delete()
        
        # Delete the classroom
        classroom.delete()
        messages.success(request, f'Lớp "{classroom.name}" đã được xóa thành công.')
        return redirect('classroom_list')
    
    return redirect('classroom_list')

@login_required
@user_passes_test(is_admin)
def classroom_add_student(request, classroom_id):
    classroom = get_object_or_404(ClassRoom, id=classroom_id)
    
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        name = request.POST.get('name')
        
        if student_id and name:
            # Kiểm tra sinh viên đã tồn tại chưa
            existing_student = Student.objects.filter(student_id=student_id).first()

            if existing_student:
                # Nếu tồn tại, báo lỗi + không thêm
                messages.error(request, f"Sinh viên ID '{existing_student.student_id}' - tên '{existing_student.name}' đã tồn tại. Không thể thêm.")
            else:
                try:
                    # Try to get existing student or create new one
                    student, created = Student.objects.get_or_create(
                        student_id=student_id,
                        defaults={
                            'name': name,
                        }
                    )
                    
                    if created:
                        # Create user account for new student
                        user = User.objects.create_user(
                            username=student_id,
                            email=f"{student_id}@example.com",
                            password='password123'  # Default password
                        )
                        student.user = user
                        student.save()
                        messages.success(request, f'Đã thêm sinh viên {name} vào lớp {classroom.name}.')
                    else:
                        # Update existing student
                        student.name = name
                        student.save()
                        messages.success(request, f'Đã cập nhật thông tin sinh viên {name}.')

                    # Thêm sinh viên vào lớp học (Many-to-Many)
                    classroom.students.add(student)
                    messages.success(request, f'Đã thêm sinh viên {name} vào lớp {classroom.name}.')
                        
                except Exception as e:
                    messages.error(request, f'Lỗi khi thêm sinh viên: {str(e)}')
        else:
            messages.error(request, 'Vui lòng điền đầy đủ thông tin sinh viên.')
            
    return redirect('classroom_import_students', classroom_id=classroom.id) 

@require_POST
@login_required
def mark_notifications_read(request):
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    return JsonResponse({'status': 'ok'})


@csrf_exempt
@login_required
def save_fcm_token(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        token = data.get('token')
        if hasattr(request.user, 'teacher'):
            teacher = request.user.teacher
            teacher.fcm_token = token
            teacher.save()

            try:
                # 👇 Sử dụng TeacherAssignment để lấy phân công mới nhất
                from core.models import TeacherAssignment
                latest_assignment = TeacherAssignment.objects.filter(teacher=teacher).order_by('-id').first()
                if latest_assignment:
                    classroom = latest_assignment.classroom
                    subject = latest_assignment.subject
                    semester = latest_assignment.semester

                    from core.fcm import send_fcm_v1
                    send_fcm_v1(
                        token,
                        "📢 Bạn vừa được phân công giảng dạy",
                        f"Lớp: {classroom.name}, Môn: {subject.name}, Kỳ: {semester.name}"
                    )
            except Exception as e:
                print("❌ Lỗi khi gửi thông báo:", str(e))

        return JsonResponse({'status': 'success'})

    return JsonResponse({'status': 'invalid'}, status=400)
